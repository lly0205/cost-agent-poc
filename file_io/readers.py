"""
File readers for case input and guidance files.
Supported formats: PDF, Excel (.xlsx/.xls), CSV, Markdown, JSON, YAML, TXT
CAD/DWG/IFC reserved for future integration.
"""
from pathlib import Path
from typing import Optional

PROJECT_ROOT = Path(__file__).parent.parent

# Max chars per file to avoid token overuse in PoC
MAX_FILE_CHARS = 8000

GUIDANCE_TASK_MAP = {
    "drawing_review": [
        "general_principles",
        "drawing_boq_consistency",
        "uncertainty_handling",
    ],
    "quantity_takeoff": [
        "general_principles",
        "decoration_quantity",
        "door_window_quantity",
        "uncertainty_handling",
        "human_review_rules",
    ],
    "boq_check": [
        "general_principles",
        "drawing_boq_consistency",
        "uncertainty_handling",
        "human_review_rules",
    ],
    "variation_claim": [
        "general_principles",
        "uncertainty_handling",
        "human_review_rules",
    ],
    "glodon_integration": ["general_principles"],
    "evaluation": ["general_principles", "human_review_rules"],
}


def read_case_files(input_dir: Path) -> dict[str, str]:
    """Read all supported files from case input directory."""
    if not input_dir.exists():
        return {}

    result = {}
    for path in sorted(input_dir.iterdir()):
        if path.is_dir():
            continue
        if path.name.startswith("."):
            continue
        content = _read_file(path)
        if content:
            result[path.name] = content

    return result


def read_guidance_files(task_type: str) -> str:
    """Load and concatenate guidance files for a given task type."""
    guidance_dir = PROJECT_ROOT / "guidance"
    file_stems = GUIDANCE_TASK_MAP.get(task_type, ["general_principles"])

    sections = []
    for stem in file_stems:
        path = guidance_dir / f"{stem}.md"
        if path.exists():
            content = path.read_text(encoding="utf-8")
            sections.append(f"### {stem}\n\n{content}")

    return "\n\n---\n\n".join(sections)


def _read_file(path: Path) -> Optional[str]:
    suffix = path.suffix.lower()

    if suffix in (".md", ".txt", ".json", ".yaml", ".yml"):
        try:
            text = path.read_text(encoding="utf-8")
            return _truncate(text, path.name)
        except Exception as e:
            return f"[读取失败: {e}]"

    if suffix == ".csv":
        return _read_csv(path)

    if suffix in (".xlsx", ".xls"):
        return _read_excel(path)

    if suffix == ".pdf":
        return _read_pdf(path)

    if suffix in (".dwg", ".dxf", ".ifc", ".rvt"):
        return (
            f"[CAD/BIM 文件：{path.name}]\n"
            "当前版本不支持自动解析此格式。\n"
            "请将关键尺寸和做法手动录入 project_instructions.md。"
        )

    return None


def _read_csv(path: Path) -> str:
    try:
        import csv
        rows = []
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i > 200:
                    rows.append("... (超过 200 行，已截断)")
                    break
                rows.append(",".join(row))
        return _truncate("\n".join(rows), path.name)
    except Exception as e:
        return f"[CSV 读取失败: {e}]"


def _read_excel(path: Path) -> str:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"[Sheet: {sheet_name}]")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > 300:
                    parts.append("... (超过 300 行，已截断)")
                    break
                row_str = "\t".join(str(c) if c is not None else "" for c in row)
                if row_str.strip():
                    parts.append(row_str)
        return _truncate("\n".join(parts), path.name)
    except ImportError:
        return f"[Excel 读取需要 openpyxl，请运行 pip install openpyxl]"
    except Exception as e:
        return f"[Excel 读取失败: {e}]"


def _read_pdf(path: Path) -> str:
    try:
        import pypdf
        reader = pypdf.PdfReader(str(path))
        texts = []
        for i, page in enumerate(reader.pages):
            if i >= 20:
                texts.append("... (超过 20 页，已截断)")
                break
            texts.append(f"[Page {i + 1}]\n{page.extract_text() or ''}")
        return _truncate("\n".join(texts), path.name)
    except ImportError:
        return (
            f"[PDF 读取需要 pypdf，请运行 pip install pypdf]\n"
            f"文件：{path.name}（请手动将关键信息录入 project_instructions.md）"
        )
    except Exception as e:
        return f"[PDF 读取失败: {e}]"


def _truncate(text: str, filename: str) -> str:
    if len(text) > MAX_FILE_CHARS:
        return text[:MAX_FILE_CHARS] + f"\n... [文件 {filename} 已截断，超过 {MAX_FILE_CHARS} 字符]"
    return text
